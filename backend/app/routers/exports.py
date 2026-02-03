"""Export router for CSV and XLSX exports of vacation requests."""
from fastapi import APIRouter, Depends, Query, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from typing import Optional
from uuid import UUID
from datetime import date, datetime
from io import StringIO, BytesIO
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models import (
    User, VacationRequest, TeamManagerAssignment, Team, UserRole
)
from app.auth import get_current_user

router = APIRouter(prefix="/export", tags=["Exports"])


async def _build_export_query(
    current_user: User,
    db: AsyncSession,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    status: Optional[str] = None,
    team_id: Optional[UUID] = None,
    user_id: Optional[UUID] = None
):
    """Build the vacation request query with role-based filtering."""
    query = select(VacationRequest).options(
        selectinload(VacationRequest.user),
        selectinload(VacationRequest.team),
        selectinload(VacationRequest.approver)
    ).where(VacationRequest.company_id == current_user.company_id)
    
    # Apply role-based filtering
    if current_user.role == UserRole.USER:
        # Users can only export their own requests
        query = query.where(VacationRequest.user_id == current_user.id)
    elif current_user.role == UserRole.MANAGER:
        # Managers can export requests from their teams
        managed_teams = await db.execute(
            select(TeamManagerAssignment.team_id).where(
                TeamManagerAssignment.user_id == current_user.id
            )
        )
        team_ids = [t[0] for t in managed_teams.fetchall()]
        if team_id and team_id in team_ids:
            team_ids = [team_id]
        if team_ids:
            query = query.where(VacationRequest.team_id.in_(team_ids))
        else:
            # Manager with no teams - return empty
            query = query.where(VacationRequest.id.is_(None))
    # Admins can see all within company
    
    # Apply filters
    if start_date:
        query = query.where(VacationRequest.end_date >= start_date)
    if end_date:
        query = query.where(VacationRequest.start_date <= end_date)
    if status:
        query = query.where(VacationRequest.status == status)
    if team_id and current_user.role == UserRole.ADMIN:
        query = query.where(VacationRequest.team_id == team_id)
    if user_id and current_user.role == UserRole.ADMIN:
        query = query.where(VacationRequest.user_id == user_id)
    
    query = query.order_by(VacationRequest.start_date.desc())
    
    return query


@router.get("/csv")
async def export_vacation_requests_csv(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    status: Optional[str] = None,
    team_id: Optional[UUID] = None,
    user_id: Optional[UUID] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Export vacation requests as CSV. Authorization enforced based on role."""
    # Build query with company isolation
    query = await _build_export_query(
        current_user, db, start_date, end_date, status, team_id, user_id
    )
    
    result = await db.execute(query)
    requests = result.scalars().all()
    
    # Generate CSV
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'ID',
        'Employee Name',
        'Employee Email',
        'Team',
        'Start Date',
        'End Date',
        'Days',
        'Type',
        'Status',
        'Reason',
        'Approver',
        'Approved At',
        'Created At'
    ])
    
    # Write data rows
    for vr in requests:
        days = (vr.end_date - vr.start_date).days + 1
        approver_name = f"{vr.approver.first_name} {vr.approver.last_name}" if vr.approver else ""
        writer.writerow([
            str(vr.id),
            f"{vr.user.first_name} {vr.user.last_name}",
            vr.user.email,
            vr.team.name if vr.team else "",
            vr.start_date.isoformat(),
            vr.end_date.isoformat(),
            str(days),
            vr.vacation_type,
            vr.status,
            vr.reason or "",
            approver_name,
            vr.approved_at.isoformat() if vr.approved_at else "",
            vr.created_at.isoformat()
        ])
    
    # Return CSV response
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=vacation_requests_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        }
    )


@router.get("/xlsx")
async def export_vacation_requests_xlsx(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    status: Optional[str] = None,
    team_id: Optional[UUID] = None,
    user_id: Optional[UUID] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Export vacation requests as XLSX. Authorization enforced based on role."""
    # Build query with company isolation
    query = await _build_export_query(
        current_user, db, start_date, end_date, status, team_id, user_id
    )
    
    result = await db.execute(query)
    requests = result.scalars().all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vacation Requests"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    
    # Write header
    headers = [
        'ID',
        'Employee Name',
        'Employee Email',
        'Team',
        'Start Date',
        'End Date',
        'Days',
        'Type',
        'Status',
        'Reason',
        'Approver',
        'Approved At',
        'Created At'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Write data rows
    for row_num, vr in enumerate(requests, 2):
        days = (vr.end_date - vr.start_date).days + 1
        approver_name = f"{vr.approver.first_name} {vr.approver.last_name}" if vr.approver else ""
        
        data = [
            str(vr.id),
            f"{vr.user.first_name} {vr.user.last_name}",
            vr.user.email,
            vr.team.name if vr.team else "",
            vr.start_date,
            vr.end_date,
            days,
            vr.vacation_type,
            vr.status,
            vr.reason or "",
            approver_name,
            vr.approved_at.isoformat() if vr.approved_at else "",
            vr.created_at.isoformat()
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            if isinstance(value, (date, datetime)):
                cell.number_format = "YYYY-MM-DD"
            cell.alignment = Alignment(horizontal="left")
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Add summary sheet
    wb.create_sheet(title="Summary")
    summary_ws = wb["Summary"]
    
    # Count by status
    status_counts = {}
    for vr in requests:
        status_counts[vr.status] = status_counts.get(vr.status, 0) + 1
    
    summary_ws.cell(row=1, column=1, value="Status")
    summary_ws.cell(row=1, column=2, value="Count")
    summary_ws["A1"].font = header_font
    summary_ws["A1"].fill = header_fill
    summary_ws["B1"].font = header_font
    summary_ws["B1"].fill = header_fill
    
    for row, (status, count) in enumerate(status_counts.items(), 2):
        summary_ws.cell(row=row, column=1, value=status)
        summary_ws.cell(row=row, column=2, value=count)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return XLSX response
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=vacation_requests_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    )
